from openpyxl import load_workbook

# Load an existing workbook
wb = load_workbook('example.xlsx')

# Select a worksheet
ws = wb['Sheet1']

# Update a cell value
ws['A1'] = 'Updated Value'

# Save the workbook
wb.save('updated_example.xlsx')
